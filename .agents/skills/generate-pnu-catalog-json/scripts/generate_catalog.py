from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[4]
RAW_DIR = ROOT / "backend" / "data" / "raw"
CATALOG_PATH = ROOT / "backend" / "data" / "course_catalog.json"
RESTRICTIONS_PATH = ROOT / "backend" / "data" / "course_restrictions.json"
DEPARTMENTS_PATH = ROOT / "frontend" / "src" / "data" / "departments.json"

RESTRICTION_FILE = "course_restriction.xlsx"
EXCLUDED_FILES = {RESTRICTION_FILE}
DAY_PATTERN = r"[월화수목금토일]"
TIME_PATTERN = r"\d{1,2}:\d{2}"
RANGE_RE = re.compile(
    rf"^(?P<day>{DAY_PATTERN})\s+(?P<start>{TIME_PATTERN})\s*-\s*"
    rf"(?P<end>{TIME_PATTERN})(?:\s+(?P<room>.+))?$"
)
DURATION_RE = re.compile(
    rf"^(?P<day>{DAY_PATTERN})\s+(?P<start>{TIME_PATTERN})\s*"
    rf"\(\s*(?P<duration>\d+)\s*\)(?:\s+(?P<room>.+))?$"
)


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    parsed = float(value)
    return int(parsed) if parsed.is_integer() else parsed


def clean_department(value: Any) -> str:
    text = clean_text(value)
    text = re.split(r"<br\s*/?>", text, maxsplit=1, flags=re.IGNORECASE)[0]
    return re.sub(r"\s+", " ", text).strip()


def parse_departments(value: Any) -> list[str]:
    raw = clean_text(value)
    if not raw:
        return []
    parts = re.split(r"\s*(?:,|;|\n|<br\s*/?>)\s*", raw, flags=re.IGNORECASE)
    return [part for part in (clean_department(item) for item in parts) if part]


def parse_capacity(value: Any) -> tuple[dict[str, int] | None, str]:
    raw = clean_text(value)
    match = re.fullmatch(r"\s*(\d+)\s*/\s*(\d+)\s*", raw)
    if not match:
        return None, raw
    return {"current": int(match.group(1)), "max": int(match.group(2))}, raw


def normalize_time(value: str) -> str:
    hour, minute = (int(part) for part in value.split(":"))
    return f"{hour:02d}:{minute:02d}"


def parse_schedules(value: Any) -> list[dict[str, Any]]:
    raw = clean_text(value)
    if not raw:
        return []
    chunks = re.split(r"\s*(?:,|\n|<br\s*/?>)\s*", raw, flags=re.IGNORECASE)
    schedules: list[dict[str, Any]] = []
    for chunk in chunks:
        if not chunk:
            continue
        match = RANGE_RE.fullmatch(chunk)
        if match:
            schedules.append(
                {
                    "day": match.group("day"),
                    "startTime": normalize_time(match.group("start")),
                    "endTime": normalize_time(match.group("end")),
                    "durationMinutes": None,
                    "room": clean_text(match.group("room")),
                }
            )
            continue
        match = DURATION_RE.fullmatch(chunk)
        if match:
            schedules.append(
                {
                    "day": match.group("day"),
                    "startTime": normalize_time(match.group("start")),
                    "endTime": None,
                    "durationMinutes": int(match.group("duration")),
                    "room": clean_text(match.group("room")),
                }
            )
            continue
        return []
    return schedules


def category_for(path: Path) -> tuple[str, int | None]:
    if path.stem == "general_required":
        return "GENERAL_REQUIRED", None
    match = re.fullmatch(r"general_elective_area_(\d)", path.stem)
    if match:
        return "GENERAL_ELECTIVE", int(match.group(1))
    raise ValueError(f"Unrecognized catalog filename: {path.name}")


def meaningful(value: Any) -> bool:
    return bool(clean_text(value))


def is_excel_data_file(path: Path) -> bool:
    return path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")


def unique_headers(row: tuple[Any, ...]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(row, 1):
        header = clean_text(value) or f"column_{index}"
        counts[header] = counts.get(header, 0) + 1
        if counts[header] > 1:
            header = f"{header}_{counts[header]}"
        headers.append(header)
    return headers


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def build_restriction_data(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    restrictions: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if "교과목번호" in {clean_text(value) for value in row}
            ),
            None,
        )
        if header_index is None:
            header_index = next(
                (
                    index
                    for index, row in enumerate(rows)
                    if sum(bool(clean_text(value)) for value in row) >= 3
                ),
                None,
            )
        if header_index is None:
            continue
        header_row = rows[header_index]
        if header_row is None:
            continue
        headers = unique_headers(header_row)
        for row_number, row in enumerate(rows[header_index + 1 :], header_index + 2):
            if not any(value not in (None, "") for value in row):
                continue
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = {
                headers[index]: clean_cell(values[index])
                for index in range(len(headers))
            }
            restrictions.append(
                {
                    "sourceFile": path.name,
                    "sheet": sheet.title,
                    "rowNumber": row_number,
                    "data": record,
                }
            )
    return restrictions


def build_course(row: tuple[Any, ...], category: str, area: int | None) -> dict[str, Any]:
    capacity, raw_capacity = parse_capacity(row[9])
    raw_target = clean_text(row[12])
    raw_offering = clean_text(row[13])
    course = {
        "category": category,
        "area": area,
        "courseType": clean_text(row[3]),
        "grade": number(row[2]),
        "courseName": clean_text(row[4]),
        "courseCode": clean_text(row[5]),
        "section": clean_text(row[6]),
        "credits": number(row[7]),
        "hours": number(row[8]),
        "capacity": capacity,
        "rawCapacity": raw_capacity,
        "professor": clean_text(row[10]),
        "rawTimeAndRoom": clean_text(row[11]),
        "schedules": parse_schedules(row[11]),
        "targetDepartments": parse_departments(row[12]),
        "offeringDepartment": clean_department(row[13]),
        "rawTargetDepartments": raw_target,
        "rawOfferingDepartment": raw_offering,
        "isCyber": meaningful(row[0]),
        "isEnglish": meaningful(row[1]),
        "notes": clean_text(row[16]),
    }
    return course


def main() -> None:
    files = sorted(
        path
        for path in RAW_DIR.glob("*.xlsx")
        if is_excel_data_file(path) and path.name not in EXCLUDED_FILES
    )
    courses: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for path in files:
        category, area = category_for(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not any(value not in (None, "") for value in row):
                    skipped.append({"file": path.name, "row": row_number, "reason": "empty row"})
                    continue
                if not clean_text(row[5]) or not clean_text(row[4]):
                    skipped.append(
                        {"file": path.name, "row": row_number, "reason": "missing course code or name"}
                    )
                    continue
                courses.append(build_course(row, category, area))

    departments = sorted(
        {department for course in courses for department in course["targetDepartments"]}
    )
    department_data = [
        {"college": None, "department": department} for department in departments
    ]

    CATALOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    RESTRICTIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    DEPARTMENTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    CATALOG_PATH.write_text(
        json.dumps(courses, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    restrictions = build_restriction_data(RAW_DIR / RESTRICTION_FILE)
    RESTRICTIONS_PATH.write_text(
        json.dumps(restrictions, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    DEPARTMENTS_PATH.write_text(
        json.dumps(department_data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    report = {
        "inputFiles": [path.name for path in files],
        "excludedFiles": sorted(EXCLUDED_FILES),
        "coursesWritten": len(courses),
        "restrictionsWritten": len(restrictions),
        "restrictionsOutput": str(RESTRICTIONS_PATH.relative_to(ROOT)),
        "departmentsWritten": len(department_data),
        "rowsSkipped": skipped,
        "timeParsingFailed": sum(
            bool(course["rawTimeAndRoom"]) and not course["schedules"] for course in courses
        ),
        "capacityParsingFailed": sum(course["capacity"] is None for course in courses),
        "departmentSource": "수강대상학과",
    }
    print(json.dumps(report, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
