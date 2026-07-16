"""Parse user-entered or uploaded courses into a fixed ``InputTimetable``.

Unlike :mod:`course_parser`, this module understands major categories. Its
output is fixed input state and is never used as an automatically recommended
course pool.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from ..models.course import Category, Course, normalize_course_category
from ..models.input_timetable import InputTimetable
from .course_parser import (
    CatalogParseError,
    _find_header_in_leading_rows,
    _cell,
    _text,
    parse_class_times,
)


class InputTimetableParseError(ValueError):
    """Raised when fixed-course input cannot form an InputTimetable."""


def _category(value: Any, default: Category = Category.MAJOR_BASIC) -> Category:
    if not _text(value):
        return default
    try:
        return normalize_course_category(value)
    except ValueError as exc:
        raise InputTimetableParseError(f"지원하지 않는 고정 과목 구분입니다: {value}") from exc


def parse_fixed_course(data: Mapping[str, Any], *, index: int = 1) -> Course:
    """Parse one API/form-style fixed course mapping."""

    code = _text(data.get("course_code") or data.get("course_id"))
    division = _text(data.get("division") or "001").zfill(3)
    course_id = _text(data.get("course_id")) or f"{code}-{division}"
    times_value = data.get("class_times")
    if isinstance(times_value, list):
        class_times = times_value
    else:
        class_times = parse_class_times(data.get("schedule") or data.get("time"))
    try:
        return Course(
            course_id=course_id,
            course_name=data.get("course_name") or data.get("name"),
            category=_category(data.get("category")),
            credit=data.get("credit"),
            division=division,
            professor=data.get("professor") or "미정",
            class_times=class_times,
        )
    except Exception as exc:
        raise InputTimetableParseError(f"고정 과목 {index}행을 해석할 수 없습니다: {exc}") from exc


def parse_fixed_courses(rows: Iterable[Mapping[str, Any]]) -> list[Course]:
    courses = [parse_fixed_course(row, index=index) for index, row in enumerate(rows, 1)]
    if not courses:
        raise InputTimetableParseError("고정 과목이 하나 이상 필요합니다.")
    return courses


def parse_input_timetable(
    rows: Iterable[Mapping[str, Any]],
    *,
    total_credit: float | None = None,
) -> InputTimetable:
    """Build and validate fixed timetable state from direct user input."""

    try:
        return InputTimetable(courses=parse_fixed_courses(rows), total_credit=total_credit)
    except InputTimetableParseError:
        raise
    except Exception as exc:
        raise InputTimetableParseError(f"고정 시간표가 유효하지 않습니다: {exc}") from exc


def parse_input_timetable_workbook(path: str | Path) -> InputTimetable:
    """Parse an uploaded fixed timetable; leading title rows are permitted."""

    source = Path(path)
    if source.suffix.lower() != ".xlsx":
        raise InputTimetableParseError(".xlsx 파일만 지원합니다.")
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise InputTimetableParseError("엑셀 파일을 열 수 없습니다.") from exc
    courses: list[Course] = []
    seen: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            try:
                header_index, columns = _find_header_in_leading_rows(values)
            except CatalogParseError as exc:
                raise InputTimetableParseError(str(exc)) from exc
            for row_index, row in enumerate(values[header_index + 1:], header_index + 2):
                code = _text(_cell(row, columns, "course_code"))
                name = _text(_cell(row, columns, "course_name"))
                division = _text(_cell(row, columns, "division")).zfill(3)
                if not code or not name:
                    continue
                class_times = parse_class_times(_cell(row, columns, "schedule"))
                if not class_times:
                    continue
                course_id = f"{code}-{division}"
                if course_id in seen:
                    continue
                mapping = {
                    "course_id": course_id,
                    "course_name": name,
                    "category": _cell(row, columns, "category"),
                    "credit": _cell(row, columns, "credit"),
                    "division": division,
                    "professor": _text(_cell(row, columns, "professor")) or "미정",
                    "class_times": class_times,
                }
                courses.append(parse_fixed_course(mapping, index=row_index))
                seen.add(course_id)
    finally:
        workbook.close()
    if not courses:
        raise InputTimetableParseError("시간 정보가 있는 고정 과목을 찾지 못했습니다.")
    try:
        return InputTimetable(courses=courses)
    except Exception as exc:
        raise InputTimetableParseError(f"고정 시간표가 유효하지 않습니다: {exc}") from exc


def parse_fixed_course_workbook(path: str | Path) -> list[Course]:
    """Compatibility helper for upload routes that store courses separately."""

    return parse_input_timetable_workbook(path).courses
