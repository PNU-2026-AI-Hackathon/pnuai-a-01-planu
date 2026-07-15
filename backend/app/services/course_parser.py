"""Parse PNU catalog workbooks into the normalized JSON used by PlaNU.

The university periodically changes header wording, so this module primarily
matches semantic header aliases and falls back to the stable column positions
of the official 17-column catalog export.
"""
# 클래스끼리 참조할 때 타입 이름을 문자열로 감싸야 NameError가 발생하지 않음
# 그러나 이것을 사용하면 문자열로 감쌀 필요가 없고, 타입 힌트가 더 깔끔해짐
from __future__ import annotations

import html
import json
import re
# pathlib.Path는 파일 경로를 객체로 다루기 위한 클래스
from pathlib import Path
# Any: 모든 타입을 허용하는 타입 힌트
# Mapping: 키-값 쌍을 가지는 매핑 타입(예: dict)
# Sequence: 순서가 있는 시퀀스 타입(예: list, tuple)
from typing import Any, Callable, Mapping, Sequence
# 엑셀 파일을 열기 위한 라이브러리
from openpyxl import load_workbook
# 프로젝트 내부 모델 가져옴
from ..models.course import Category, ClassTime, Course, Day, normalize_course_category

# 수강편람에서 특정 정보가 어느 열에 있는지 명시
CATALOG_POSITION = {
    "category": 3,
    "course_name": 4,
    "course_code": 5,
    "division": 6,
    "credit": 7,
    "professor": 10,
    "schedule": 11,
}

# 수강 편람마다 열 이름이 바뀔 수 있기에 aliases를 통해 예외 방지
HEADER_ALIASES = {
    "category": ("교과구분", "이수구분", "교과목구분"),
    "course_name": ("교과목명", "과목명", "교과목명(정확한명칭)"),
    "course_code": ("교과목번호", "과목번호", "학수번호"),
    "division": ("분반",),
    "credit": ("학점",),
    "professor": ("담당교수", "교수명", "교수"),
    "schedule": ("시간/강의실", "수업시간/강의실", "강의시간/강의실"),
}

# 요일을 enum으로 변환
DAY_MAP = {
    "월": Day.MON, "화": Day.TUE, "수": Day.WED, "목": Day.THU, "금": Day.FRI,
    "MON": Day.MON, "TUE": Day.TUE, "WED": Day.WED, "THU": Day.THU, "FRI": Day.FRI,
}
# 수업 시간과 강의실을 추출하는 정규식
_MEETING_RE = re.compile(
    # 매칭되는 문자가 있다면 day라는 이름에 저장(뒤에 공백 있어도 됨)
    r"(?P<day>월|화|수|목|금|MON|TUE|WED|THU|FRI)\s*"
    # 매칭되는 문자가 있다면 start라는 이름에 저장
    # 숫자 1~2자리, 콜론, 숫자 2자리 형식으로 시작 시간 추출
    r"(?P<start>\d{1,2}:\d{2})\s*"
    # 매칭되는 문자가 있다면 end라는 이름에 저장
    # 시작 시간과 종료 시간 사이에 - 또는 ~가 올 수 있음
    r"(?:[-~]\s*(?P<end>\d{1,2}:\d{2})|\(\s*(?P<duration>\d+)\s*\))\s*"
    # 매칭되는 문자가 있다면 classroom이라는 이름에 저장
    # 강의실 정보는 월~금과 시간 정보 사이에 위치하며, 쉼표, 세미콜론, 슬래시 등으로 구분될 수 있음
    # .*?는 가능한 한 적게 매칭하도록 하여(다음 요일 or 시간이 나오거나 문자열 끝이 나오면 종료), 강의실 정보가 끝나는 위치를 정확히 찾음
    r"(?P<classroom>.*?)(?=(?:\s*[,;/]\s*|\s+)(?:월|화|수|목|금|MON|TUE|WED|THU|FRI)\s*\d{1,2}:\d{2}|$)",
    re.IGNORECASE,
)

# 커스텀 예외로 엑셀을 열 수 없을 때 발생
class CatalogParseError(ValueError):
    """Raised when an xlsx file is not a recognizable course catalog."""

# 001 같은 값을 1로 읽는 것을 방지
# 엑셀에서 읽은 모든 값을 문자열로 변환하고, None은 빈 문자열로 변환
# 3.0 같이 정수로 표현 가능한 실수는 int로 변환하여 문자열로 변환
def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

# 열 이름을 정규화하여 비교
def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s()（）_\-]", "", _text(value)).lower()

# 특정 행에서 열 이름을 찾아 열 번호를 반환
def _header_columns(row: Sequence[Any]) -> dict[str, int]:
    normalized = [_normalized_header(value) for value in row]
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for index, value in enumerate(normalized):
            #aliases 중 하나라도 value에 포함되어 있으면 result에 field와 index를 저장
            if any(_normalized_header(alias) in value for alias in aliases):
                result[field] = index
                break
    return result

_REQUIRED_CATALOG_FIELDS = {
    "course_name", "course_code", "division", "credit", "schedule"
}


def _catalog_columns_from_header(row: Sequence[Any]) -> dict[str, int]:
    """Resolve one known catalog header row, with positional fallback."""

    columns = _header_columns(row)
    if not _REQUIRED_CATALOG_FIELDS.issubset(columns) and len(row) >= 12:
        columns = dict(CATALOG_POSITION)
    if not _REQUIRED_CATALOG_FIELDS.issubset(columns):
        raise CatalogParseError("필수 열(교과목명, 교과목번호, 분반, 학점, 시간/강의실)을 찾을 수 없습니다.")
    return columns


def _find_header_in_leading_rows(
    rows: Sequence[Sequence[Any]],
    *,
    score: Callable[[Sequence[Any]], int] | None = None,
    limit: int = 30,
) -> tuple[int, dict[str, int]]:
    """Search leading rows only for documents that may contain title rows."""

    if not rows:
        raise CatalogParseError("엑셀 파일이 비어 있습니다.")
    if score is not None:
        index = max(
            range(min(limit, len(rows))),
            key=lambda row_index: score(rows[row_index]),
        )
        return index, {}

    candidates = [(_header_columns(row), index) for index, row in enumerate(rows[:limit])]
    columns, index = max(candidates, key=lambda item: len(item[0]))
    if not _REQUIRED_CATALOG_FIELDS.issubset(columns):
        # Broken headers cannot be scored semantically. In that case the
        # table header is normally the densest leading row.
        index = max(
            range(min(limit, len(rows))),
            key=lambda row_index: sum(value is not None for value in rows[row_index]),
        )
        columns = _catalog_columns_from_header(rows[index])
    return index, columns

# 시간 포맷을 HH:MM으로 변환

def _clock(value: str) -> str:
    hour, minute = (int(part) for part in value.split(":"))
    return f"{hour:02d}:{minute:02d}"

# 수업 길이만 주어지는 경우엔 시작 시간에 분을 더해 종료 시간을 계산
def _add_minutes(value: str, minutes: int) -> str:
    hour, minute = (int(part) for part in value.split(":"))
    total = hour * 60 + minute + minutes
    if not 0 <= total < 24 * 60:
        raise CatalogParseError(f"잘못된 수업 시간입니다: {value}({minutes})")
    return f"{total // 60:02d}:{total % 60:02d}"

# 건물번호를 추출하는 함수
def _building_code(classroom: str) -> str:
    # PNU room strings include forms such as 609-313 and 양산M03-3B20.
    match = re.search(r"(?:^|\s)([가-힣A-Za-z]*\d{3}|[가-힣A-Za-z]+M\d{2})(?=-|\s|$)", classroom)
    if match:
        return match.group(1)
    token = classroom.split("-", 1)[0].strip().split()[-1:] or ["미정"]
    return token[0] or "미정"

# 수업 시간 문자열을 class_time 객체로 변환
def parse_class_times(value: Any) -> list[ClassTime]:
    """Parse PNU's mixed ``09:00-10:15`` / ``09:00(75)`` notation."""
    # 불필요한 공백과 HTML 엔티티를 제거하고, <br> 태그를 쉼표로 대체
    raw = html.unescape(_text(value))
    raw = re.sub(r"<br\s*/?>", ", ", raw, flags=re.IGNORECASE)
    # 엑셀이나 웹에서 사용되는 특수문자 제거
    raw = raw.replace("，", ",").replace("～", "~").replace("–", "-")

    meetings: list[ClassTime] = []
    seen: set[tuple[Day, str, str]] = set()
    for match in _MEETING_RE.finditer(raw):
        classroom = match.group("classroom").strip(" ,;/") or "미정"
        start = _clock(match.group("start"))
        end = _clock(match.group("end")) if match.group("end") else _add_minutes(start, int(match.group("duration")))
        day = DAY_MAP[match.group("day").upper()]
        key = (day, start, end)
        if key in seen:
            continue
        meetings.append(ClassTime(
            day=day,
            start=start,
            end=end,
            classroom=classroom,
            building_code=_building_code(classroom),
        ))
        seen.add(key)
    return meetings

# 특정 행에서 열 값을 추출하는 함수로 열 번호가 없거나 범위를 벗어나면 None을 반환
def _cell(row: Sequence[Any], columns: Mapping[str, int], field: str) -> Any:
    index = columns.get(field, -1)
    return row[index] if 0 <= index < len(row) else None


def parse_catalog_workbook(
    path: str | Path,
    category: Category | str,
    *,
    area: int | None = None,
) -> list[Course]:
    """Parse an internal general-course catalog whose first row is its header."""

    try:
        category = normalize_course_category(category)
    except ValueError as exc:
        raise CatalogParseError(str(exc)) from exc
    if category not in (Category.GENERAL_REQUIRED, Category.GENERAL_ELECTIVE):
        raise CatalogParseError("내부 수강편람 파서는 교양필수/교양선택만 처리합니다.")
    if category is Category.GENERAL_ELECTIVE and area is None:
        raise CatalogParseError("교양선택 수강편람에는 영역 번호가 필요합니다.")
    source = Path(path)
    # Path 객체를 사용하고 있기에 suffix가 파일 확장자를 의미한다.
    if source.suffix.lower() != ".xlsx":
        raise CatalogParseError(".xlsx 파일만 지원합니다.")
    if not source.is_file():
        raise FileNotFoundError(source)
    courses: list[Course] = []
    seen: set[str] = set()
    try:
        # read_only=True 이기에 엑셀 파일을 가볍게 열 수 있음.
        # data_only=True 이기에 수식이 아닌 계산된 값만 읽음(셀 객체가 아닌 셀 값만 가져온다는 의미).
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise CatalogParseError("엑셀 파일을 열 수 없습니다.") from exc
    try:
        for sheet in workbook.worksheets:
            # 각 행이 튜플로 이루어진 리스트를 생성.
            # 그렇기에 다른 함수들은 매개변수로 Sequence[Any] 타입을 받음.
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            # 내부 수강편람은 첫 행이 헤더라는 계약을 갖는다. 여기서는
            # 선행 행 탐색을 하지 않으며, 깨진 헤더만 위치 fallback 한다.
            columns = _header_columns(rows[0])
            if not _REQUIRED_CATALOG_FIELDS.issubset(columns):
                columns = _catalog_columns_from_header(rows[0])
            for row in rows[1:]:
                code = _text(_cell(row, columns, "course_code"))
                name = _text(_cell(row, columns, "course_name"))
                raw_division = _text(_cell(row, columns, "division"))
                if not code or not name or not raw_division:
                    continue
                # 원본 값을 확인한 뒤에만 3자리 분반으로 정규화한다.
                division = raw_division.zfill(3)
                if not division.isdigit() or division == "000":
                    continue
                try:
                    credit = float(_cell(row, columns, "credit"))
                except (TypeError, ValueError):
                    continue
                class_times = parse_class_times(_cell(row, columns, "schedule"))
                if not class_times:  # Online/TBA rows cannot be scheduled safely.
                    continue
                course_id = f"{code}-{division}"
                if course_id in seen:
                    continue
                courses.append(Course(
                    course_id=course_id,
                    course_name=name,
                    category=category,
                    area=area,
                    credit=credit,
                    division=division,
                    professor=_text(_cell(row, columns, "professor")) or "미정",
                    class_times=class_times,
                ))
                seen.add(course_id)
    finally:
        workbook.close()
    return courses

# 수강 지도 파일을 읽는 함수
def parse_restrictions(path: str | Path) -> tuple[list[dict[str, Any]], list[str]]:
    """Preserve restriction rows and derive a unique department list."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    departments: set[str] = set()
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            # 제한 파일은 제목/안내 행을 허용하므로 이 경로에서만 탐색한다.
            header_index, _ = _find_header_in_leading_rows(
                rows,
                score=lambda row: sum(value is not None for value in row),
                limit=20,
            )
            headers = [_text(v) or f"column_{i + 1}" for i, v in enumerate(rows[header_index])]
            for row in rows[header_index + 1:]:
                if not any(_text(v) for v in row):
                    continue
                record = {headers[i]: _text(value) for i, value in enumerate(row) if i < len(headers) and _text(value)}
                records.append(record)
                # Official layout: offering department and restricted department.
                for index in (1, 7):
                    if index < len(row) and _text(row[index]):
                        departments.add(_text(row[index]).split("<br", 1)[0])
    finally:
        workbook.close()
    return records, sorted(departments)

# JSON 파일을 안전하게 쓰는 함수. 임시 파일을 만들고, 성공하면 덮어씀.
def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    # ensure_ascii=False: 한글이 깨지지 않도록 함
    # indent=2: JSON 파일을 보기 좋게 들여쓰기
    temporary.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(path)

# 과거 raw 엑셀 여러 개를 읽어 processed JSON을 만들던 수동 변환 함수.
# 서버 기본 데이터 로딩은 course_loader.load_default_catalogs를 사용한다.
def build_processed_data(
    raw_dir: str | Path,
    processed_dir: str | Path,
) -> dict[str, int]:
    """Build legacy processed JSON files from raw xlsx files."""

    raw, output = Path(raw_dir), Path(processed_dir)
    required = parse_catalog_workbook(raw / "general_required.xlsx", Category.GENERAL_REQUIRED)
    electives: list[Course] = []
    for area in range(1, 8):
        electives.extend(parse_catalog_workbook(
            raw / f"general_elective_area_{area}.xlsx",
            Category.GENERAL_ELECTIVE,
            area=area,
        ))
    restriction_path = raw / "course_restriction_rules.xlsx"
    if not restriction_path.exists():
        restriction_path = raw / "course_restriction.xlsx"
    restrictions, departments = parse_restrictions(restriction_path)
    _write_json(output / "general_required_courses.json", [c.model_dump(mode="json") for c in required])
    _write_json(output / "general_elective_courses.json", [c.model_dump(mode="json") for c in electives])
    _write_json(output / "course_restrictions.json", restrictions)
    _write_json(output / "department_list.json", departments)
    # 각 데이터의 개수를 반환
    return {"general_required": len(required), "general_elective": len(electives),
            "restrictions": len(restrictions), "departments": len(departments)}


# Backwards-friendly name for focused unit tests and upload parsing.
parse_course_file = parse_catalog_workbook


def parse_default_catalogs(catalog_path: str | Path) -> dict[str, list[Course]]:
    """Load default server catalog data from the generated JSON file."""

    from .course_loader import load_default_catalogs

    return load_default_catalogs(catalog_path)
